from guiPrCh import *
from openpyxl import *
import sys
from PyQt5.QtWidgets import QFileDialog, QTableWidgetItem, QMessageBox


all_arts = []
double_arts = []
del_rows = []
count_check = 0  #Счётчик проверенных файлов
count_change = {}  #Словарь изменённых цен
file1_path = ''
file2_path = ''
file1 = None  # файл с новыми ценами
file2 = None  # файл со старыми ценами
nf_art_column = None  # колонка с артиклем в новом файле
of_art_column = None  # колонка с артиклем
sheet_nf = 0; wb1 = 0; ws1 = 0; sheet_of = 0; wb2 = 0; ws2 = 0; mr1 = 0; mr2 = 0; mc2 = 0
start_row_nf = None
start_row_of = None
count_del = 0
del_arts = []
not_find = 0
missing_arts = []


class MyWin(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        QtWidgets.QWidget.__init__(self, parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setStyleSheet("background-color: Beige;")



        # Здесь прописываем события нажатия на кнопку
        self.ui.result_Button.clicked.connect(self.pr_check)
        self.ui.del_excess.clicked.connect(self.del_ex)
        self.ui.del_repeat_items.clicked.connect(self.del_rep)
        self.ui.del_repeat_single.clicked.connect(self.del_rep_single)
        self.ui.del_repeat_single_sum.clicked.connect(lambda: self.del_rep_single(True))
        self.ui.n_f_Button.clicked.connect(lambda : self.select_file(self.ui.n_f_path, self.ui.n_f_table,
                                                                     int(self.ui.nf_sheet_value.text()
                                                                     )))
        self.ui.o_f_Button.clicked.connect(lambda : self.select_file(self.ui.o_f_path, self.ui.o_f_table,
                                                                     int(self.ui.of_sheet_value.text())))


    #Далее идут функции
#------------------------------------------------------------------#
    # Главная функция сравнения цен
    def pr_check(self):
        if not self.all_check('all'):
            return
        global wb1, count_check, all_arts, double_arts, sheet_nf, wb1, ws1, sheet_of, wb2, ws2, mr1, mr2, mc2, file1, \
            file2, nf_art_column, of_art_column, \
            start_row_nf, start_row_of, count_del, del_arts, not_find, missing_arts
        self.table()
        nf_pr_column = int(self.ui.n_f_price_value.text())  # колнкой с ценой в новом файле
        of_pr_column = int(self.ui.o_f_price_value.text())  # колонка со старой ценой
        of_npr_column = int(self.ui.o_f_new_price_value.text())  # колонка для записи новой цены
        for i in range(start_row_nf, mr1 + 1):
            art = str(ws1.cell(row=i, column=nf_art_column).value).strip()
            if art == None or art == '' or art == 'None':
                continue
            if art in all_arts:
                double_arts.append(art)
                continue
            else:
                all_arts.append(art)
                find = 0
            for j in range(start_row_of, mr2 + 1):
                if count_check >= mr2:
                    break
                if art == str(ws2.cell(row=j, column=of_art_column).value).strip():
                    find = 1
                    if str(ws1.cell(row=i, column=nf_pr_column).value).strip() != str(ws2.cell(row=j, column=of_pr_column).value).strip():
                        # Заносим арт в словарь
                        count_change[art] = {ws2.cell(row=j, column=of_npr_column).value: str(ws1.cell(row=i, column=nf_pr_column).value)}
                        ws2.cell(row=j, column=of_npr_column).value = str(ws1.cell(row=i, column=nf_pr_column).value).strip()
                        count_check += 1
            if not find:
                missing_arts.append(art)
        wb2.save(str(file2))
        self.results()


    #Выводим текст результатов
    def results(self):
        global count_check, count_change, all_arts, double_arts
        self.ui.result_text.clear()
        text = f'Поменяли {count_check} значений!\nДублей в исходном файле - {len(double_arts)}'
        self.ui.result_text.setText(text)
        if len(missing_arts) > 0:
            self.ui.result_text.append(f"Не нашли {len(missing_arts)} значений")
        count_check = 0
        #Заполняем таблицу результатами
        if len(missing_arts) > len(double_arts):
            mis_or_dbl = len(missing_arts)
        else:
            mis_or_dbl = len(double_arts)
        table_rows = len(count_change) + mis_or_dbl + 3 #количество рядов в
        # таблице
        self.ui.result_table.setColumnCount(3)
        self.ui.result_table.setRowCount(table_rows)
        font = QtGui.QFont()
        font.setBold(True)
        row = 0
        item = str('Дубли')
        self.ui.result_table.setItem(row, 0, QTableWidgetItem(item))
        self.ui.result_table.item(row, 0).setBackground(QtGui.QColor(143, 188, 143))
        self.ui.result_table.item(row, 0).setFont(font)
        item = str('Не нашли')
        self.ui.result_table.setItem(row, 1, QTableWidgetItem(item))
        self.ui.result_table.item(row, 1).setBackground(QtGui.QColor(143, 188, 143))
        self.ui.result_table.item(row, 1).setFont(font)
        row += 1
        for i in range(0, mis_or_dbl):
            if i < len(double_arts):
                item = str(double_arts[i])
                self.ui.result_table.setItem(row, 0, QTableWidgetItem(item))
                self.ui.result_table.item(row, 0).setBackground(QtGui.QColor(255, 255, 0))
            if i < len(missing_arts):
                item = str(missing_arts[i])
                self.ui.result_table.setItem(row, 1, QTableWidgetItem(item))
                self.ui.result_table.item(row, 1).setBackground(QtGui.QColor(255, 255, 0))
            row += 1

        item = str('Код')
        self.ui.result_table.setItem(row, 0, QTableWidgetItem(item))
        self.ui.result_table.item(row, 0).setBackground(QtGui.QColor(143, 188, 143))
        self.ui.result_table.item(row, 0).setFont(font)
        item = str('Старое значение')
        self.ui.result_table.setItem(row, 1, QTableWidgetItem(item))
        self.ui.result_table.item(row, 1).setBackground(QtGui.QColor(143, 188, 143))
        self.ui.result_table.item(row, 1).setFont(font)
        item = str('Новое значение')
        self.ui.result_table.setItem(row, 2, QTableWidgetItem(item))
        self.ui.result_table.item(row, 2).setBackground(QtGui.QColor(143, 188, 143))
        self.ui.result_table.item(row, 2).setFont(font)
        row += 1
        for key, value in count_change.items():
            for inkey, invalue in value.items():
                item = str(key)
                self.ui.result_table.setItem(row, 0, QTableWidgetItem(item))
                item = str(inkey)
                self.ui.result_table.setItem(row, 1, QTableWidgetItem(item))
                item = str(invalue)
                self.ui.result_table.setItem(row, 2, QTableWidgetItem(item))
            row += 1
        row += 1
        self.ui.result_table.resizeColumnsToContents()
        self.ui.result_table.resizeRowsToContents()
        count_change.clear()
        double_arts.clear()
        all_arts.clear()
        missing_arts.clear()


    #Функция выбора файлов и заполнения примеров таблицы
    def select_file(self, entry, table, sheet):
        entry.setText(QFileDialog.getOpenFileName()[0])
        #Заполняем первые строки из выбранной таблицы
        if not entry.text():
            return
        wb = load_workbook(entry.text())
        ws = wb.worksheets[sheet - 1]
        mc = ws.max_column
        mr = ws.max_row
        if mr > 15:
            mr = 15
        table.setRowCount(mr)
        table.setColumnCount(mc)
        for i in range(1, mr + 1):
            for j in range(1, mc + 1):
                item = str(ws.cell(i, j).value)
                table.setRowHeight(i - 1, 5)
                table.setItem(i - 1, j - 1, QTableWidgetItem(item))


    #Функция отправляет на проверку каждое из заполняемых полей
    def all_check(self, check_type):
        if not self.ui.n_f_path.text():
            QMessageBox.about(self, 'Не все поля заполнены', 'Укажите файл с данными')
            return False
        if not self.check(self.ui.n_f_art_value.text()):
            QMessageBox.about(self, 'Не все поля заполнены', 'Укажите Код')
            return False
        if check_type == 'for del_single':
            return True
        if check_type == 'for del_single+':
            if not self.check(self.ui.n_f_price_value.text()):
                QMessageBox.about(self, 'Не все поля заполнены', 'Укажите Значение')
                return False
            return True
        if not self.ui.o_f_path.text():
            QMessageBox.about(self, 'Не все поля заполнены', 'Укажите файл с данными')
            return False
        if not self.check(self.ui.o_f_art_value.text()):
            QMessageBox.about(self, 'Не все поля заполнены', 'Укажите Код')
            return False
        if check_type == 'for del':
            return True
        if not self.check(self.ui.n_f_price_value.text()):
            QMessageBox.about(self, 'Не все поля заполнены', 'Укажите Значение')
            return False
        if not self.check(self.ui.o_f_price_value.text()):
            QMessageBox.about(self, 'Не все поля заполнены', 'Укажите Значение')
            return False
        if not self.check(self.ui.o_f_new_price_value.text()):
            QMessageBox.about(self, 'Не все поля заполнены', 'Укажите куда вводить новое Значение')
            return False
        return True


    # Функция проверяет каждое из заполняемых полей
    def check(self, val):
        try:
            int(val)
            return True
        except ValueError:
            return False


    #Просто очищаем ячейки в лишних строках
    def del_ex(self):
        if not self.all_check('for del'):
            return
        global wb1, ws1, wb2, ws2, mr1, mr2, file2, nf_art_column, of_art_column, \
            start_row_nf, start_row_of, count_del, del_arts, not_find, missing_arts, del_rows
        self.table()
        for i in range(start_row_of, mr2 + 1):
            art2 = str(ws2.cell(row=i, column=of_art_column).value).strip()
            if art2 == None or art2 == '' or art2 == 'None':
                continue
            find = 0
            for x in range(start_row_nf, mr1 + 1):
                if art2 == str(ws1.cell(row=x, column=nf_art_column).value).strip():
                    find = 1
                    break
            if find == 1:
                continue
            del_arts.append(art2)
            del_rows.append(i)
            count_del += 1
        for i in range(start_row_nf, mr1 + 1):
            art1 = str(ws1.cell(row=i, column=nf_art_column).value).strip()
            if art1 == None or art1 == '' or art1 == 'None':
                continue
            find = 0
            for x in range(start_row_of, mr2 + 1):
                if art1 == str(ws2.cell(row=x, column=of_art_column).value).strip():
                    find = 1
                    break
            if find == 1:
                continue
            missing_arts.append(art1)
            not_find += 1
        if len(del_rows) > 0:
            for val in sorted(del_rows)[::-1]:
                ws2.delete_rows(val)
        wb2.save(str(file2))
        self.del_result()


    #Удаление строк с одинаковым значением
    def del_rep(self):
        if not self.all_check('for del'):
            return
        global wb1, ws1, wb2, ws2, mr1, mr2, file2, nf_art_column, of_art_column, \
            start_row_nf, start_row_of, count_del, del_arts, not_find, missing_arts
        self.table()
        #Находим и очищаем дублирующиеся строки
        for i in range(start_row_nf, mr1 + 1):
            art1 = str(ws1.cell(row=i, column=nf_art_column).value).strip()
            if art1 == None or art1 == '' or art1 == 'None':
                continue
            find = 0
            for x in range(start_row_of, mr2 + 1):
                if count_del >= mr2:
                    break
                if art1 == str(ws2.cell(row=x, column=of_art_column).value).strip():
                    find = 1
                    if art1 not in del_arts:
                        del_arts.append(art1)
                    if x not in del_rows:
                        del_rows.append(x)
                        count_del += 1
            if find == 0:
                missing_arts.append(art1)
                not_find += 1
        if len(del_rows) > 0:
            for val in  sorted(del_rows)[::-1]:
                ws2.delete_rows(val)
        wb2.save(str(file2))
        self.del_result()


    def table(self, all=True):  # открываем таблицы
        global sheet_nf, wb1, ws1, sheet_of, wb2, ws2, mr1, mr2, mc2, file1, file2, nf_art_column, of_art_column, \
            start_row_nf, start_row_of, count_del, del_arts, not_find, missing_arts, del_rows

        self.ui.result_text.clear()
        self.ui.result_table.clear()
        count_del = 0
        if len(del_arts) > 0:
            del_arts.clear()
        if len(del_rows) > 0:
            del_rows.clear()
        if len(missing_arts) > 0:
            missing_arts.clear()
        not_find = 0
        # открытие исходного файла Excel
        file1 = self.ui.n_f_path.text()  # файл с новыми ценами
        nf_art_column = int(self.ui.n_f_art_value.text())  # колнка с артиклем в новом файле
        sheet_nf = int(self.ui.nf_sheet_value.text()) - 1
        wb1 = load_workbook(file1)
        ws1 = wb1.worksheets[sheet_nf]
        mr1 = ws1.max_row
        start_row_nf = int(self.ui.nf_start_row_value.text())
        if not all:
            return
        # открытие целевого файла Excel
        file2 = self.ui.o_f_path.text()  # файл со старыми ценами
        of_art_column = int(self.ui.o_f_art_value.text())  # колонка с артиклем
        sheet_of = int(self.ui.of_sheet_value.text()) - 1
        wb2 = load_workbook(file2)
        ws2 = wb2.worksheets[sheet_of]
        mr2 = ws2.max_row
        start_row_of = int(self.ui.of_start_row_value.text())


    def del_result(self):
        global sheet_nf, wb1, ws1, sheet_of, wb2, ws2, mr1, mr2, mc2, file1, file2, nf_art_column, of_art_column, \
            start_row_nf, start_row_of, count_del, del_arts, not_find, missing_arts
        if len(missing_arts) > len(del_arts):
            table_rows = len(missing_arts) + 1  # количество рядов в таблице
        else:
            table_rows = len(del_arts) + 1  # количество рядов в таблице
        self.ui.result_table.setColumnCount(3)
        self.ui.result_table.setRowCount(table_rows)
        row = 0
        if len(missing_arts) > 0:
            font = QtGui.QFont()
            font.setBold(True)
            item = str('Не нашли')
            self.ui.result_table.setItem(row, 0, QTableWidgetItem(item))
            self.ui.result_table.item(row, 0).setBackground(QtGui.QColor(143, 188, 143))
            self.ui.result_table.item(row, 0).setFont(font)
            self.ui.result_text.append(f"Не нашли {not_find} строк!")
            row += 1
            for i in missing_arts:
                item = str(i)
                self.ui.result_table.setItem(row, 0, QTableWidgetItem(item))
                self.ui.result_table.item(row, 0).setBackground(QtGui.QColor(255, 255, 0))
                self.ui.result_text.append(f"{item}")
                row += 1
        row += 1
        self.ui.result_text.append(f"Очистили {count_del} строк!")
        if len(del_arts) > 0:
            row = 0
            font = QtGui.QFont()
            font.setBold(True)
            item = str('Очистили')
            self.ui.result_table.setItem(row, 2, QTableWidgetItem(item))
            self.ui.result_table.item(row, 2).setBackground(QtGui.QColor(143, 188, 143))
            self.ui.result_table.item(row, 2).setFont(font)
            row += 1
            for i in del_arts:
                item = str(i)
                self.ui.result_table.setItem(row, 2, QTableWidgetItem(item))
                self.ui.result_text.append(f"{item}")
                row += 1

    # Удаление строк с одинаковым значением
    def del_rep_single(self, sum_values=False):
            type_check = 'for del_single+' if sum_values else 'for del_single'
            if not self.all_check(type_check):
                return
            if sum_values:
                nf_pr_column = int(self.ui.n_f_price_value.text())
            global wb1, ws1, mr1, file1, nf_art_column, start_row_nf, count_del, del_arts
            self.table(all=False)
            # Находим и очищаем дублирующиеся строки
            for i in range(start_row_nf, mr1 + 1):
                art1 = str(ws1.cell(row=i, column=nf_art_column).value).strip()
                if art1 == None or art1 == '' or art1 == 'None':
                    continue
                for x in range(mr1, i, - 1):
                    if art1 == str(ws1.cell(row=x, column=nf_art_column).value).strip():
                        if sum_values:
                            founded_val = float(ws1.cell(row=i, column=nf_pr_column).value)
                            ws1.cell(row=i, column=nf_pr_column).value = str(float(ws1.cell(row=x, column=nf_pr_column).value) + founded_val)
                        ws1.delete_rows(x)
                        del_arts.append(art1)
                        count_del += 1
            wb1.save(str(file1))
            self.del_result()


if __name__=="__main__":
    app = QtWidgets.QApplication(sys.argv)
    myapp = MyWin()
    myapp.show()
    sys.exit(app.exec_())
